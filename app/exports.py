"""Excel (.xlsx) and branded PDF exports. openpyxl + reportlab are required
(both present); a CSV fallback is provided if openpyxl is ever unavailable."""
import io
from datetime import datetime, timezone

from flask import send_file, Response

from config import Config

BASE = Config.__module__  # unused; keep import light
_CHARCOAL = "3C3C3C"
_GREY = "7E8080"
_LOGO = None


def _logo_path():
    from pathlib import Path
    p = Path(__file__).resolve().parent / "static" / "img" / "logo-mark.png"
    return str(p) if p.exists() else None


def xlsx_response(filename, sheet_name, headers, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception:
        return _csv_response(filename.replace(".xlsx", ".csv"), headers, rows)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    head_fill = PatternFill("solid", fgColor=_CHARCOAL)
    head_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(bottom=thin)
    # Title row
    ws.append([f"{Config.ORG_NAME} · {sheet_name}"])
    ws["A1"].font = Font(bold=True, size=13, name="Arial", color=_CHARCOAL)
    ws.append([f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M} UTC"])
    ws["A2"].font = Font(size=9, color=_GREY, name="Arial")
    ws.append([])
    ws.append(headers)
    hr = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=hr, column=c)
        cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for r in rows:
        ws.append(list(r))
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).border = border
    for c in range(1, len(headers) + 1):
        width = max([len(str(headers[c - 1]))] + [len(str(r[c - 1])) for r in rows[:200] if c - 1 < len(r)] + [8])
        ws.column_dimensions[chr(64 + c) if c <= 26 else "AA"].width = min(width + 3, 48)
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _csv_response(filename, headers, rows):
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(list(r))
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


def pdf_report_response(filename, title, meta, sections):
    """sections: list of (heading, headers, rows). meta: dict with generated_by, filters."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                        Spacer, Image)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except Exception:
        return Response("PDF engine unavailable", status=501)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm, title=title)
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontName="Helvetica-Bold",
                       fontSize=18, textColor=colors.HexColor("#" + _CHARCOAL), spaceAfter=2)
    sub = ParagraphStyle("s", parent=styles["Normal"], fontName="Helvetica",
                         fontSize=8.5, textColor=colors.HexColor("#" + _GREY))
    sect = ParagraphStyle("sec", parent=styles["Heading2"], fontName="Helvetica-Bold",
                          fontSize=12, textColor=colors.HexColor("#" + _CHARCOAL), spaceBefore=14, spaceAfter=6)
    flow = []
    lp = _logo_path()
    header_bits = []
    if lp:
        img = Image(lp, width=13 * mm, height=13 * mm)
        header_bits.append(img)
    txt = [Paragraph(f"{Config.ORG_NAME} · {title}", h),
           Paragraph(f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M} UTC · "
                     f"by {meta.get('generated_by','—')}", sub)]
    if meta.get("filters"):
        txt.append(Paragraph("Filters: " + meta["filters"], sub))
    if header_bits:
        flow.append(Table([[header_bits[0], txt]], colWidths=[16 * mm, None],
                          style=TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")])))
    else:
        flow.extend(txt)
    flow.append(Spacer(1, 6))
    flow.append(Table([[""]], colWidths=[doc.width],
                      style=TableStyle([("LINEBELOW", (0, 0), (-1, -1), 0.6, colors.HexColor("#DDDDDD"))])))

    for heading, headers, rows in sections:
        flow.append(Paragraph(heading, sect))
        data = [headers] + [list(r) for r in rows]
        t = Table(data, repeatRows=1, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + _CHARCOAL)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F2")]),
            ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E2E2")),
            ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ]))
        flow.append(t)

    def _footer(canvas, d):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#" + _GREY))
        canvas.drawString(16 * mm, 10 * mm, f"{Config.ORG_NAME} · T-CAP Enterprise Control Suite")
        canvas.drawRightString(A4[0] - 16 * mm, 10 * mm, f"Page {d.page}")
        canvas.restoreState()

    doc.build(flow, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/pdf")
