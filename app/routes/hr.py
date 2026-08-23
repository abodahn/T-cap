"""HR module — employee directory + Excel importer.

Phase 1: a searchable/paginated staff directory and an in-app importer that
loads the HR master spreadsheet (Arabic or English headers) into the DB.
Confidential fields (salary, insurance wage, bank details, advances) are only
rendered to users with hr_manage. Employee data is never seeded into git — it
enters the database only through this importer at runtime.
"""
from flask import (Blueprint, render_template, request, redirect, url_for,
                   abort, flash)

from app.db import get_db, utcnow, log_audit, next_ref
from app.auth import permission_required, login_required, current_user, user_can

bp = Blueprint("hr", __name__, url_prefix="/hr")

# HR service request types (key -> EN label; AR comes from i18n hrmod_<key>).
HR_MODULES = [
    ("services", "Leave & HR Services"), ("absence", "Absence Notification"),
    ("no_show", "No-Show Follow-up"), ("exit", "Exit / Resignation"),
    ("probation", "Probation Completion"), ("contracts", "Contract Renewal"),
    ("appraisals", "Performance Appraisal"), ("benefits", "Benefits & Medical"),
    ("training", "Training Request"), ("policies", "Policy / Document Request"),
    ("recognition", "Recognition"), ("other", "Other HR Request"),
]
HR_MODULE_KEYS = {k for k, _ in HR_MODULES}
CASE_PRIORITIES = ["Low", "Normal", "High"]
# Workflow. Terminal = no outgoing transitions.
CASE_STATUSES = ["Submitted", "In Review", "More Info", "Approved", "Rejected",
                 "Completed", "Cancelled"]
CASE_OPEN = {"Submitted", "In Review", "More Info", "Approved"}
_TRANSITIONS = {
    "Submitted": {"In Review", "Cancelled"},
    "In Review": {"Approved", "Rejected", "More Info", "Cancelled"},
    "More Info": {"In Review", "Cancelled"},
    "Approved": {"Completed"},
    "Rejected": set(), "Completed": set(), "Cancelled": set(),
}
CASE_PAGE = 25

# Spreadsheet header (Arabic master or English) -> employees column.
HEADER_MAP = {
    "الكود": "employee_no", "code": "employee_no", "employee_no": "employee_no", "emp_no": "employee_no", "id": "employee_no",
    "الاسم بالعربية": "name_ar", "name_ar": "name_ar", "arabic name": "name_ar",
    "الاسم بالإنجليزية": "name_en", "name_en": "name_en", "english name": "name_en", "name": "name_en", "الاسم": "name_en",
    "الإدارة": "department", "department": "department", "dept": "department",
    "القسم": "section", "section": "section",
    "المسمى الوظيفي": "job_title", "job_title": "job_title", "title": "job_title",
    "job": "job_title", "position": "job_title", "الوظيفة": "job_title",
    "email": "email", "e-mail": "email", "mail": "email", "البريد الإلكتروني": "email", "البريد": "email",
    "mob": "phone", "mobile": "phone", "phone": "phone", "tel": "phone",
    "الهاتف": "phone", "الموبايل": "phone", "الجوال": "phone",
    "الموقع": "location", "location": "location", "address": "location", "العنوان": "location",
    "المرتب الأساسي": "basic_salary", "basic_salary": "basic_salary", "basic": "basic_salary",
    "الأجر التأميني": "insurance_wage", "insurance_wage": "insurance_wage",
    "كود البنك": "bank_code", "bank_code": "bank_code",
    "اسم البنك": "bank_name", "bank_name": "bank_name", "bank": "bank_name",
    "الحساب البنكي": "bank_account", "bank_account": "bank_account", "account": "bank_account",
    "يمكن الدفع؟": "payable", "payable": "payable",
    "رصيد السلفة": "advance_balance", "advance_balance": "advance_balance",
    "قسط السلفة": "advance_installment", "advance_installment": "advance_installment",
    "الحالة / المشاكل": "notes", "الحالة": "notes", "notes": "notes", "status": "notes",
}
_NUMERIC = {"basic_salary", "insurance_wage", "advance_balance", "advance_installment"}
# Columns hidden from the list and gated behind hr_manage on the detail view.
CONFIDENTIAL = ("basic_salary", "insurance_wage", "bank_code", "bank_name",
                "bank_account", "advance_balance", "advance_installment")
PAGE = 50


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _cell(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _map_header(row):
    """Return {col_index: field} for a candidate header row."""
    idx = {}
    for i, h in enumerate(row):
        if h is None:
            continue
        field = HEADER_MAP.get(str(h).strip()) or HEADER_MAP.get(str(h).strip().lower())
        if field and i not in idx:
            idx[i] = field
    return idx


def import_workbook(db, ws):
    """Upsert employees from a worksheet. The header row is auto-detected (the
    first row that maps to a name or code), so junk/title rows above it are
    tolerated. Rows are keyed by employee_no when present, otherwise by name.
    Duplicate source columns for one field keep the first non-empty value.
    Returns (inserted, updated, skipped)."""
    idx = None
    data_rows = []
    for row in ws.iter_rows(values_only=True):
        if idx is None:
            cand = _map_header(row)
            if "employee_no" in cand.values() or "name_en" in cand.values():
                idx = cand
            continue
        data_rows.append(row)
    if not idx:
        raise ValueError("no_recognisable_header")
    key_field = "employee_no" if "employee_no" in idx.values() else "name_en"

    ins = upd = skip = 0
    now = utcnow()
    for row in data_rows:
        data = {}
        for i, field in idx.items():
            val = row[i] if i < len(row) else None
            val = _num(val) if field in _NUMERIC else _cell(val)
            if val is not None and data.get(field) in (None, ""):   # first non-empty wins
                data[field] = val
        key = data.get(key_field)
        if not key:
            skip += 1
            continue
        cols = [c for c in data if c != key_field]
        existing = db.execute(f"SELECT id FROM employees WHERE {key_field}=?", (key,)).fetchone()
        if existing:
            if cols:
                sets = ", ".join(f"{c}=?" for c in cols) + ", updated_at=?"
                db.execute(f"UPDATE employees SET {sets} WHERE {key_field}=?",
                           [data[c] for c in cols] + [now, key])
            upd += 1
        else:
            allcols = [key_field] + cols + ["created_at", "updated_at"]
            ph = ",".join(["?"] * len(allcols))
            db.execute(f"INSERT INTO employees({','.join(allcols)}) VALUES({ph})",
                       [key] + [data[c] for c in cols] + [now, now])
            ins += 1
    return ins, upd, skip


@bp.route("/")
@login_required
def overview():
    """HR module home — role-aware KPI tiles, module cards, recent activity."""
    db = get_db()
    if not (user_can("hr_request") or user_can("hr_view_all")):
        abort(403)
    u = current_user()
    staff = user_can("hr_view_all")
    OPEN = "('Submitted','In Review','More Info','Approved')"
    k = {}
    if staff:
        k["headcount"] = db.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        k["open_cases"] = db.execute(f"SELECT COUNT(*) FROM hr_cases WHERE status IN {OPEN}").fetchone()[0]
        k["pending"] = db.execute("SELECT COUNT(*) FROM hr_cases WHERE status IN ('Submitted','In Review','More Info')").fetchone()[0]
        k["appraisals"] = db.execute("SELECT COUNT(*) FROM hr_appraisals").fetchone()[0]
    k["my_open"] = db.execute(f"SELECT COUNT(*) FROM hr_cases WHERE (created_by=? OR requester=?) AND status IN {OPEN}",
                              (u["id"], u["full_name"])).fetchone()[0]
    k["policies"] = db.execute("SELECT COUNT(*) FROM hr_policies WHERE active=1").fetchone()[0]
    acked = {r[0] for r in db.execute("SELECT policy_id FROM hr_policy_acks WHERE user_id=?", (u["id"],)).fetchall()}
    k["policies_todo"] = max(0, k["policies"] - len(acked))
    announcements = db.execute("SELECT * FROM hr_announcements ORDER BY pinned DESC, created_at DESC LIMIT 3").fetchall()
    if staff:
        recent = db.execute(f"SELECT * FROM hr_cases ORDER BY created_at DESC LIMIT 6").fetchall()
    else:
        recent = db.execute("SELECT * FROM hr_cases WHERE created_by=? OR requester=? ORDER BY created_at DESC LIMIT 6",
                            (u["id"], u["full_name"])).fetchall()
    return render_template("hr/overview.html", k=k, staff=staff, announcements=announcements,
                           recent=recent, module_label=_module_label,
                           can_manage=user_can("hr_manage"))


@bp.route("/directory")
@permission_required("hr_view_all")
def index():
    db = get_db()
    q = (request.args.get("q") or "").strip()
    dept = request.args.get("dept") or ""
    where, p = "1=1", []
    if q:
        where += " AND (name_en LIKE ? OR name_ar LIKE ? OR employee_no LIKE ?)"; p += [f"%{q}%"] * 3
    if dept:
        where += " AND department=?"; p.append(dept)
    total = db.execute("SELECT COUNT(*) FROM employees WHERE " + where, p).fetchone()[0]
    try:
        page = max(1, int(request.args.get("page") or 1))
    except ValueError:
        page = 1
    pages = max(1, (total + PAGE - 1) // PAGE)
    rows = db.execute("SELECT * FROM employees WHERE " + where +
                      " ORDER BY name_en, employee_no LIMIT ? OFFSET ?",
                      p + [PAGE, (page - 1) * PAGE]).fetchall()
    depts = [r[0] for r in db.execute("SELECT DISTINCT department FROM employees WHERE department IS NOT NULL AND department<>'' ORDER BY department").fetchall()]
    return render_template("hr/list.html", employees=rows, q=q, dept=dept, depts=depts,
                           total=total, shown=len(rows), page=page, pages=pages,
                           can_manage=user_can("hr_manage"))


@bp.route("/employee/<int:eid>")
@permission_required("hr_view_all")
def view(eid):
    db = get_db()
    e = db.execute("SELECT * FROM employees WHERE id=?", (eid,)).fetchone()
    if not e:
        abort(404)
    mgr = db.execute("SELECT name_en,employee_no FROM employees WHERE id=?", (e["manager_id"],)).fetchone() if e["manager_id"] else None
    return render_template("hr/detail.html", e=e, mgr=mgr,
                           can_manage=user_can("hr_manage"), confidential=CONFIDENTIAL)


@bp.route("/import", methods=["GET", "POST"])
@permission_required("hr_manage")
def import_data():
    db = get_db()
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename.lower().endswith((".xlsx", ".xlsm")):
            flash("hr_import_bad_file")
            return redirect(url_for("hr.import_data"))
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ins, upd, skip = import_workbook(db, wb.active)
            db.commit()
            log_audit(current_user()["username"], "hr_import", f"+{ins} ~{upd} skip{skip}")
            flash(f"hr_import_done:{ins}:{upd}:{skip}")
        except ValueError as ex:
            flash("hr_import_no_code" if "header" in str(ex) else "hr_import_error")
        except Exception:
            flash("hr_import_error")
        return redirect(url_for("hr.index"))
    return render_template("hr/import.html")


# ===========================================================================
#  HR service cases — request → review → decision, with an activity trail.
#  Visibility mirrors ITSM: HR staff (hr_view_all) see all; everyone else
#  sees only the cases they raised.
# ===========================================================================

def _case_event(db, cid, kind, summary, detail=""):
    u = current_user()
    db.execute("""INSERT INTO hr_case_events(case_id,actor,kind,summary,detail,created_at)
                  VALUES(?,?,?,?,?,?)""",
               (cid, u["full_name"] if u else "system", kind, summary, detail, utcnow()))


def _case_owns(c):
    u = current_user()
    if not u:
        return False
    if c["created_by"] is not None:
        return c["created_by"] == u["id"]
    return (c["requester"] or "") == (u["full_name"] or "")


def _can_view_case(c):
    return user_can("hr_view_all") or _case_owns(c)


def _module_label(key):
    return dict(HR_MODULES).get(key, key)


@bp.route("/cases")
@login_required
def cases():
    db = get_db()
    if not (user_can("hr_request") or user_can("hr_view_all")):
        abort(403)
    u = current_user()
    staff = user_can("hr_view_all")
    status = request.args.get("status") or ""
    scope = request.args.get("scope") or ""      # staff: '', 'open', 'mine-queue'
    q = (request.args.get("q") or "").strip()
    where, p = "1=1", []
    if not staff:                                # requesters see only their own
        where += " AND (created_by=? OR requester=?)"; p += [u["id"], u["full_name"]]
    if status:
        where += " AND status=?"; p.append(status)
    if scope == "open":
        where += " AND status IN ('Submitted','In Review','More Info','Approved')"
    elif scope == "queue" and staff:
        where += " AND assignee=?"; p.append(u["full_name"])
    if q:
        where += " AND (subject LIKE ? OR ref LIKE ? OR requester LIKE ?)"; p += [f"%{q}%"] * 3
    total = db.execute("SELECT COUNT(*) FROM hr_cases WHERE " + where, p).fetchone()[0]
    try:
        page = max(1, int(request.args.get("page") or 1))
    except ValueError:
        page = 1
    pages = max(1, (total + CASE_PAGE - 1) // CASE_PAGE)
    rows = db.execute("SELECT * FROM hr_cases WHERE " + where +
                      " ORDER BY (status IN ('Completed','Rejected','Cancelled')), "
                      "CASE priority WHEN 'High' THEN 0 WHEN 'Normal' THEN 1 ELSE 2 END, created_at DESC"
                      " LIMIT ? OFFSET ?", p + [CASE_PAGE, (page - 1) * CASE_PAGE]).fetchall()
    open_mine = db.execute("SELECT COUNT(*) FROM hr_cases WHERE (created_by=? OR requester=?) AND status IN ('Submitted','In Review','More Info','Approved')",
                           (u["id"], u["full_name"])).fetchone()[0]
    return render_template("hr/cases.html", cases=rows, statuses=CASE_STATUSES,
                           f_status=status, scope=scope, q=q, staff=staff,
                           total=total, shown=len(rows), page=page, pages=pages,
                           open_mine=open_mine, module_label=_module_label)


@bp.route("/cases/new", methods=["GET", "POST"])
@login_required
def case_new():
    db = get_db()
    if not user_can("hr_request"):
        abort(403)
    u = current_user()
    if request.method == "POST":
        module = request.form.get("module") or "other"
        if module not in HR_MODULE_KEYS:
            module = "other"
        subject = (request.form.get("subject") or "").strip()
        if not subject:
            flash("hr_case_subject_required")
            return redirect(url_for("hr.case_new"))
        priority = request.form.get("priority") if request.form.get("priority") in CASE_PRIORITIES else "Normal"
        ref = next_ref(db, "hr_cases", "ref", "HR", year=True)
        db.execute("""INSERT INTO hr_cases(ref,module,subject,description,requester,created_by,
                      department,status,priority,created_at,updated_at)
                      VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                   (ref, module, subject, (request.form.get("description") or "").strip(),
                    u["full_name"], u["id"], u["department"], "Submitted", priority,
                    utcnow(), utcnow()))
        cid = db.execute("SELECT id FROM hr_cases WHERE ref=?", (ref,)).fetchone()["id"]
        _case_event(db, cid, "system", "Request submitted", _module_label(module))
        db.commit()
        log_audit(u["username"], "hr_case_create", ref)
        return redirect(url_for("hr.case_view", ref=ref))
    return render_template("hr/case_form.html", modules=HR_MODULES, priorities=CASE_PRIORITIES)


@bp.route("/cases/<ref>")
@login_required
def case_view(ref):
    db = get_db()
    c = db.execute("SELECT * FROM hr_cases WHERE ref=?", (ref,)).fetchone()
    if not c:
        abort(404)
    if not _can_view_case(c):
        abort(403)
    events = db.execute("SELECT * FROM hr_case_events WHERE case_id=? ORDER BY created_at, id", (c["id"],)).fetchall()
    can_decide = user_can("hr_manage")
    can_act = can_decide or _case_owns(c)
    nexts = sorted(_TRANSITIONS.get(c["status"], set()))
    return render_template("hr/case_detail.html", c=c, events=events, can_decide=can_decide,
                           can_act=can_act, is_owner=_case_owns(c), nexts=nexts,
                           module_label=_module_label, open_=c["status"] in CASE_OPEN)


@bp.route("/cases/<ref>/action", methods=["POST"])
@login_required
def case_action(ref):
    db = get_db()
    c = db.execute("SELECT * FROM hr_cases WHERE ref=?", (ref,)).fetchone()
    if not c:
        abort(404)
    if not _can_view_case(c):
        abort(403)
    u = current_user()
    a = request.form.get("action")
    now = utcnow()
    staff = user_can("hr_manage")

    if a == "comment":
        body = (request.form.get("comment") or "").strip()
        if body:
            _case_event(db, c["id"], "comment", body)

    elif a == "cancel":
        # Requester (or HR) may cancel their own case while it's still open.
        if c["status"] in _TRANSITIONS and "Cancelled" in _TRANSITIONS[c["status"]]:
            if not (staff or _case_owns(c)):
                abort(403)
            db.execute("UPDATE hr_cases SET status='Cancelled', updated_at=?, closed_at=? WHERE id=?", (now, now, c["id"]))
            _case_event(db, c["id"], "status", "Cancelled", (request.form.get("reason") or "").strip())

    elif a == "assign" and staff:
        who = request.form.get("assignee") or ""
        db.execute("UPDATE hr_cases SET assignee=?, updated_at=? WHERE id=?", (who, now, c["id"]))
        _case_event(db, c["id"], "assign", f"Assigned to {who or 'unassigned'}")

    elif a == "transition" and staff:
        to = request.form.get("to_status") or ""
        if to in _TRANSITIONS.get(c["status"], set()):
            # No self-approval.
            if to == "Approved" and _case_owns(c):
                flash("hr_case_no_self_approve")
                return redirect(url_for("hr.case_view", ref=ref))
            reason = (request.form.get("reason") or "").strip()
            closed = now if to in ("Rejected", "Completed", "Cancelled") else None
            decision = to if to in ("Approved", "Rejected") else c["decision"]
            db.execute("""UPDATE hr_cases SET status=?, decision=?, decision_reason=?,
                          updated_at=?, closed_at=COALESCE(?, closed_at) WHERE id=?""",
                       (to, decision, reason or c["decision_reason"], now, closed, c["id"]))
            _case_event(db, c["id"], "status", f"Status → {to}", reason)
    else:
        abort(403)

    db.commit()
    log_audit(u["username"], f"hr_case_{a}", ref)
    return redirect(url_for("hr.case_view", ref=ref) + "#activity")


# ===========================================================================
#  Payroll runs — monthly per-employee payslips. Confidential: hr_manage only.
# ===========================================================================

# Payroll sheet header (EN or AR) -> payroll_runs column.
PAYROLL_MAP = {
    "name": "name", "الاسم": "name", "employee": "name",
    "code": "employee_no", "الكود": "employee_no", "employee_no": "employee_no",
    "role": "role", "job": "role", "الوظيفة": "role",
    "year": "year", "السنة": "year", "month": "month", "الشهر": "month",
    "basic": "basic", "الأساسي": "basic", "الأجر الأساسي": "basic",
    "allowances": "allowances", "البدلات": "allowances",
    "ot hrs": "ot_hours", "ot hours": "ot_hours", "overtime hrs": "ot_hours",
    "ot rate": "ot_rate", "gross": "gross", "الإجمالي": "gross",
    "deductions": "deductions", "الخصومات": "deductions",
    "absence ded": "absence_ded", "tax ded": "tax_ded",
    "net salary": "net", "net": "net", "الصافي": "net",
    "status": "status", "الحالة": "status",
}
_PAY_NUM = {"basic", "allowances", "ot_hours", "ot_rate", "gross", "deductions",
            "absence_ded", "tax_ded", "net"}
PAYROLL_STATUSES = ["Draft", "Approved", "Paid"]
PAY_MONEY = ["basic", "allowances", "gross", "deductions", "absence_ded", "tax_ded", "net"]


def import_payroll(db, ws):
    """Parse a payroll worksheet and replace the affected periods wholesale.
    Header row is auto-detected (the row containing a 'Name'/'الاسم' cell).
    Returns (rows_imported, [periods])."""
    header_idx = None
    idx = {}
    data_rows = []
    for row in ws.iter_rows(values_only=True):
        if header_idx is None:
            cells = {(str(v).strip().lower() if v is not None else "") for v in row}
            if "name" in cells or "الاسم" in cells:
                for i, h in enumerate(row):
                    if h is None:
                        continue
                    f = PAYROLL_MAP.get(str(h).strip().lower()) or PAYROLL_MAP.get(str(h).strip())
                    if f:
                        idx[i] = f
                header_idx = True
            continue
        data_rows.append(row)
    if not idx or "name" not in idx.values():
        raise ValueError("no_payroll_header")

    recs = []
    for row in data_rows:
        rec = {}
        for i, f in idx.items():
            v = row[i] if i < len(row) else None
            rec[f] = _num(v) if f in _PAY_NUM else _cell(v)
        if not rec.get("name"):
            continue
        rec["period"] = f"{rec.get('year') or ''}-{rec.get('month') or ''}".strip("-") or "unspecified"
        recs.append(rec)
    if not recs:
        raise ValueError("no_payroll_rows")

    periods = sorted({r["period"] for r in recs})
    now = utcnow()
    for per in periods:                          # re-import replaces a period
        db.execute("DELETE FROM payroll_runs WHERE period=?", (per,))
    cols = ["period", "year", "month", "employee_no", "name", "role", "basic",
            "allowances", "ot_hours", "ot_rate", "gross", "deductions",
            "absence_ded", "tax_ded", "net", "status", "created_at", "updated_at"]
    ph = ",".join(["?"] * len(cols))
    for r in recs:
        db.execute(f"INSERT INTO payroll_runs({','.join(cols)}) VALUES({ph})",
                   [r.get("period"), r.get("year"), r.get("month"), r.get("employee_no"),
                    r.get("name"), r.get("role"), r.get("basic"), r.get("allowances"),
                    r.get("ot_hours"), r.get("ot_rate"), r.get("gross"), r.get("deductions"),
                    r.get("absence_ded"), r.get("tax_ded"), r.get("net"),
                    r.get("status") or "Draft", now, now])
    return len(recs), periods


@bp.route("/payroll")
@permission_required("hr_manage")
def payroll():
    db = get_db()
    periods = [r[0] for r in db.execute("SELECT DISTINCT period FROM payroll_runs ORDER BY period DESC").fetchall()]
    period = request.args.get("period") or (periods[0] if periods else "")
    rows = db.execute("SELECT * FROM payroll_runs WHERE period=? ORDER BY name", (period,)).fetchall() if period else []
    totals = {c: sum((r[c] or 0) for r in rows) for c in PAY_MONEY}
    return render_template("hr/payroll.html", periods=periods, period=period, rows=rows,
                           totals=totals, money=PAY_MONEY)


@bp.route("/payroll/<int:pid>")
@permission_required("hr_manage")
def payslip(pid):
    db = get_db()
    r = db.execute("SELECT * FROM payroll_runs WHERE id=?", (pid,)).fetchone()
    if not r:
        abort(404)
    return render_template("hr/payslip.html", r=r)


@bp.route("/payroll/import", methods=["GET", "POST"])
@permission_required("hr_manage")
def payroll_import():
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename.lower().endswith((".xlsx", ".xlsm")):
            flash("hr_import_bad_file")
            return redirect(url_for("hr.payroll_import"))
        db = get_db()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            n, periods = import_payroll(db, wb.active)
            db.commit()
            log_audit(current_user()["username"], "hr_payroll_import", f"{n} rows / {','.join(periods)}")
            flash(f"hr_payroll_done:{n}:{len(periods)}")
            return redirect(url_for("hr.payroll", period=periods[-1] if periods else ""))
        except ValueError as ex:
            flash("hr_payroll_no_header" if "header" in str(ex) else "hr_payroll_empty")
        except Exception:
            flash("hr_import_error")
        return redirect(url_for("hr.payroll_import"))
    return render_template("hr/payroll_import.html")


# ===========================================================================
#  Announcements — HR broadcasts, read by all staff, posted by hr_manage.
# ===========================================================================
@bp.route("/announcements")
@login_required
def announcements():
    db = get_db()
    rows = db.execute("SELECT * FROM hr_announcements ORDER BY pinned DESC, created_at DESC").fetchall()
    return render_template("hr/announcements.html", items=rows, can_manage=user_can("hr_manage"))


@bp.route("/announcements/new", methods=["POST"])
@permission_required("hr_manage")
def announcement_new():
    db = get_db()
    title = (request.form.get("title") or "").strip()
    body = (request.form.get("body") or "").strip()
    if title and body:
        db.execute("INSERT INTO hr_announcements(title,body,author,pinned,created_at) VALUES(?,?,?,?,?)",
                   (title, body, current_user()["full_name"], 1 if request.form.get("pinned") else 0, utcnow()))
        db.commit()
        log_audit(current_user()["username"], "hr_announce", title[:60])
    return redirect(url_for("hr.announcements"))


@bp.route("/announcements/<int:aid>/delete", methods=["POST"])
@permission_required("hr_manage")
def announcement_delete(aid):
    db = get_db()
    db.execute("DELETE FROM hr_announcements WHERE id=?", (aid,))
    db.commit()
    return redirect(url_for("hr.announcements"))


# ===========================================================================
#  Policies & acknowledgements — staff read + acknowledge; HR tracks compliance.
# ===========================================================================
@bp.route("/policies")
@login_required
def policies():
    db = get_db()
    uid = current_user()["id"]
    rows = db.execute("SELECT * FROM hr_policies WHERE active=1 ORDER BY created_at DESC").fetchall()
    acked = {r[0] for r in db.execute("SELECT policy_id FROM hr_policy_acks WHERE user_id=?", (uid,)).fetchall()}
    total_users = db.execute("SELECT COUNT(*) FROM users WHERE is_active=1").fetchone()[0]
    counts = {r["policy_id"]: r["c"] for r in db.execute("SELECT policy_id, COUNT(*) AS c FROM hr_policy_acks GROUP BY policy_id").fetchall()}
    return render_template("hr/policies.html", items=rows, acked=acked, counts=counts,
                           total_users=total_users, can_manage=user_can("hr_manage"))


@bp.route("/policies/new", methods=["POST"])
@permission_required("hr_manage")
def policy_new():
    db = get_db()
    title = (request.form.get("title") or "").strip()
    body = (request.form.get("body") or "").strip()
    if title and body:
        db.execute("INSERT INTO hr_policies(title,body,version,active,author,created_at) VALUES(?,?,?,1,?,?)",
                   (title, body, (request.form.get("version") or "1.0").strip(),
                    current_user()["full_name"], utcnow()))
        db.commit()
        log_audit(current_user()["username"], "hr_policy_new", title[:60])
    return redirect(url_for("hr.policies"))


@bp.route("/policies/<int:pid>/ack", methods=["POST"])
@login_required
def policy_ack(pid):
    db = get_db()
    u = current_user()
    if db.execute("SELECT 1 FROM hr_policies WHERE id=? AND active=1", (pid,)).fetchone():
        db.execute("INSERT OR IGNORE INTO hr_policy_acks(policy_id,user_id,user_name,acknowledged_at) VALUES(?,?,?,?)",
                   (pid, u["id"], u["full_name"], utcnow()))
        db.commit()
    return redirect(url_for("hr.policies"))


@bp.route("/policies/<int:pid>")
@permission_required("hr_view_all")
def policy_detail(pid):
    db = get_db()
    pol = db.execute("SELECT * FROM hr_policies WHERE id=?", (pid,)).fetchone()
    if not pol:
        abort(404)
    acks = db.execute("SELECT * FROM hr_policy_acks WHERE policy_id=? ORDER BY acknowledged_at DESC", (pid,)).fetchall()
    total_users = db.execute("SELECT COUNT(*) FROM users WHERE is_active=1").fetchone()[0]
    return render_template("hr/policy_detail.html", pol=pol, acks=acks, total_users=total_users)


# ===========================================================================
#  Appraisals — performance reviews per employee (HR-managed).
# ===========================================================================
APPRAISAL_STATUSES = ["Draft", "Shared", "Acknowledged"]


@bp.route("/appraisals")
@permission_required("hr_view_all")
def appraisals():
    db = get_db()
    q = (request.args.get("q") or "").strip()
    where, p = "1=1", []
    if q:
        where += " AND (employee_name LIKE ? OR employee_no LIKE ? OR period LIKE ?)"; p += [f"%{q}%"] * 3
    rows = db.execute("SELECT * FROM hr_appraisals WHERE " + where + " ORDER BY created_at DESC", p).fetchall()
    return render_template("hr/appraisals.html", items=rows, q=q, can_manage=user_can("hr_manage"))


@bp.route("/appraisals/new", methods=["GET", "POST"])
@permission_required("hr_manage")
def appraisal_new():
    db = get_db()
    if request.method == "POST":
        name = (request.form.get("employee_name") or "").strip()
        if not name:
            flash("hr_appraisal_name_required")
            return redirect(url_for("hr.appraisal_new"))
        try:
            rating = min(5, max(1, int(request.form.get("rating") or 3)))
        except ValueError:
            rating = 3
        db.execute("""INSERT INTO hr_appraisals(employee_no,employee_name,period,reviewer,rating,
                      strengths,improvements,goals,status,created_at,updated_at)
                      VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                   (request.form.get("employee_no") or "", name,
                    (request.form.get("period") or "").strip(), current_user()["full_name"], rating,
                    (request.form.get("strengths") or "").strip(),
                    (request.form.get("improvements") or "").strip(),
                    (request.form.get("goals") or "").strip(), "Draft", utcnow(), utcnow()))
        aid = db.execute("SELECT MAX(id) AS m FROM hr_appraisals").fetchone()["m"]
        db.commit()
        log_audit(current_user()["username"], "hr_appraisal_new", name)
        return redirect(url_for("hr.appraisal_detail", aid=aid))
    emps = db.execute("SELECT employee_no,name_en FROM employees ORDER BY name_en").fetchall()
    return render_template("hr/appraisal_form.html", emps=emps, statuses=APPRAISAL_STATUSES)


@bp.route("/appraisals/<int:aid>")
@permission_required("hr_view_all")
def appraisal_detail(aid):
    db = get_db()
    a = db.execute("SELECT * FROM hr_appraisals WHERE id=?", (aid,)).fetchone()
    if not a:
        abort(404)
    return render_template("hr/appraisal_detail.html", a=a, statuses=APPRAISAL_STATUSES,
                           can_manage=user_can("hr_manage"))


@bp.route("/appraisals/<int:aid>/status", methods=["POST"])
@permission_required("hr_manage")
def appraisal_status(aid):
    db = get_db()
    st = request.form.get("status")
    if st in APPRAISAL_STATUSES:
        db.execute("UPDATE hr_appraisals SET status=?, updated_at=? WHERE id=?", (st, utcnow(), aid))
        db.commit()
    return redirect(url_for("hr.appraisal_detail", aid=aid))


# ===========================================================================
#  Self-service ("My HR") + Leave management
# ===========================================================================

def _my_employee(u):
    """Best-effort link from a user account to an employee record:
    explicit user_id, else matching email, else matching name."""
    if not u:
        return None
    db = get_db()
    row = db.execute("SELECT * FROM employees WHERE user_id=?", (u["id"],)).fetchone()
    if row:
        return row
    if u["email"]:
        row = db.execute("SELECT * FROM employees WHERE LOWER(email)=LOWER(?)", (u["email"],)).fetchone()
        if row:
            return row
    return db.execute("SELECT * FROM employees WHERE name_en=? OR name_ar=?",
                      (u["full_name"], u["full_name"])).fetchone()


def _leave_balance(employee_id):
    """Per-type entitlement / used / remaining for the current year."""
    db = get_db()
    year = utcnow()[:4]
    out = []
    for t in db.execute("SELECT * FROM leave_types WHERE active=1 ORDER BY id").fetchall():
        used = db.execute("""SELECT COALESCE(SUM(days),0) AS d FROM leave_requests
                             WHERE employee_id=? AND leave_type_id=? AND status='Approved'
                             AND substr(start_date,1,4)=?""",
                          (employee_id, t["id"], year)).fetchone()["d"] or 0
        ent = t["days_per_year"] or 0
        out.append({"type": t, "entitlement": ent, "used": used, "remaining": ent - used})
    return out


@bp.route("/me")
@login_required
def me():
    db = get_db()
    u = current_user()
    emp = _my_employee(u)
    payslips = leave_rows = balances = []
    if emp:
        payslips = db.execute("""SELECT * FROM payroll_runs WHERE name=? OR (employee_no IS NOT NULL AND employee_no=?)
                                 ORDER BY period DESC LIMIT 6""",
                              (emp["name_en"] or emp["name_ar"], emp["employee_no"])).fetchall()
        balances = _leave_balance(emp["id"])
        leave_rows = db.execute("SELECT lr.*, lt.name AS type_name FROM leave_requests lr "
                                "LEFT JOIN leave_types lt ON lt.id=lr.leave_type_id "
                                "WHERE lr.employee_id=? ORDER BY lr.created_at DESC LIMIT 5", (emp["id"],)).fetchall()
    my_cases = db.execute(f"SELECT * FROM hr_cases WHERE created_by=? OR requester=? ORDER BY created_at DESC LIMIT 5",
                          (u["id"], u["full_name"])).fetchall()
    pol_todo = db.execute("""SELECT COUNT(*) AS c FROM hr_policies p WHERE p.active=1
                             AND NOT EXISTS(SELECT 1 FROM hr_policy_acks a WHERE a.policy_id=p.id AND a.user_id=?)""",
                          (u["id"],)).fetchone()["c"]
    return render_template("hr/my_hr.html", emp=emp, payslips=payslips, balances=balances,
                           leave_rows=leave_rows, my_cases=my_cases, pol_todo=pol_todo,
                           can_manage=user_can("hr_manage"), module_label=_module_label)


LEAVE_STATUSES = ["Pending", "Approved", "Rejected", "Cancelled"]


def _days_between(start, end):
    from datetime import date
    s = date.fromisoformat(start); e = date.fromisoformat(end)
    return (e - s).days + 1


@bp.route("/leave")
@login_required
def leave():
    db = get_db()
    u = current_user()
    staff = user_can("hr_view_all")
    status = request.args.get("status") or ""
    where, p = "1=1", []
    if not staff:
        emp = _my_employee(u)
        where += " AND (lr.created_by=? OR lr.employee_id=?)"; p += [u["id"], emp["id"] if emp else -1]
    if status:
        where += " AND lr.status=?"; p.append(status)
    rows = db.execute("SELECT lr.*, lt.name AS type_name FROM leave_requests lr "
                      "LEFT JOIN leave_types lt ON lt.id=lr.leave_type_id WHERE " + where +
                      " ORDER BY (lr.status='Pending') DESC, lr.start_date DESC", p).fetchall()
    pending = db.execute("SELECT COUNT(*) AS c FROM leave_requests WHERE status='Pending'").fetchone()["c"] if staff else 0
    balances = []
    if not staff:
        emp = _my_employee(u)
        if emp:
            balances = _leave_balance(emp["id"])
    return render_template("hr/leave.html", rows=rows, staff=staff, statuses=LEAVE_STATUSES,
                           f_status=status, pending=pending, balances=balances)


@bp.route("/leave/new", methods=["GET", "POST"])
@login_required
def leave_new():
    db = get_db()
    u = current_user()
    emp = _my_employee(u)
    types = db.execute("SELECT * FROM leave_types WHERE active=1 ORDER BY id").fetchall()
    if request.method == "POST":
        try:
            lt = int(request.form.get("leave_type_id") or 0)
        except ValueError:
            lt = 0
        start = (request.form.get("start_date") or "").strip()
        end = (request.form.get("end_date") or "").strip()
        try:
            days = _days_between(start, end)
        except (ValueError, TypeError):
            days = 0
        if not lt or days < 1:
            flash("leave_invalid")
            return redirect(url_for("hr.leave_new"))
        db.execute("""INSERT INTO leave_requests(employee_id,employee_name,leave_type_id,start_date,
                      end_date,days,reason,status,created_by,created_at,updated_at)
                      VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                   (emp["id"] if emp else None, (emp["name_en"] if emp else u["full_name"]),
                    lt, start, end, days, (request.form.get("reason") or "").strip(),
                    "Pending", u["id"], utcnow(), utcnow()))
        db.commit()
        log_audit(u["username"], "leave_request", f"{days}d")
        return redirect(url_for("hr.leave"))
    balances = _leave_balance(emp["id"]) if emp else []
    return render_template("hr/leave_form.html", types=types, balances=balances, emp=emp)


@bp.route("/leave/<int:lid>")
@login_required
def leave_view(lid):
    db = get_db()
    r = db.execute("SELECT lr.*, lt.name AS type_name FROM leave_requests lr "
                   "LEFT JOIN leave_types lt ON lt.id=lr.leave_type_id WHERE lr.id=?", (lid,)).fetchone()
    if not r:
        abort(404)
    u = current_user()
    is_owner = (r["created_by"] == u["id"])
    if not (user_can("hr_view_all") or is_owner):
        abort(403)
    return render_template("hr/leave_detail.html", r=r, is_owner=is_owner,
                           can_decide=user_can("hr_manage"))


@bp.route("/leave/<int:lid>/action", methods=["POST"])
@login_required
def leave_action(lid):
    db = get_db()
    r = db.execute("SELECT * FROM leave_requests WHERE id=?", (lid,)).fetchone()
    if not r:
        abort(404)
    u = current_user()
    is_owner = (r["created_by"] == u["id"])
    if not (user_can("hr_view_all") or is_owner):
        abort(403)
    a = request.form.get("action")
    now = utcnow()
    if a in ("approve", "reject") and user_can("hr_manage") and r["status"] == "Pending":
        st = "Approved" if a == "approve" else "Rejected"
        db.execute("UPDATE leave_requests SET status=?, decided_by=?, decided_at=?, updated_at=? WHERE id=?",
                   (st, u["full_name"], now, now, lid))
        log_audit(u["username"], f"leave_{a}", str(lid))
    elif a == "cancel" and (is_owner or user_can("hr_manage")) and r["status"] in ("Pending", "Approved"):
        db.execute("UPDATE leave_requests SET status='Cancelled', updated_at=? WHERE id=?", (now, lid))
        log_audit(u["username"], "leave_cancel", str(lid))
    else:
        abort(403)
    db.commit()
    return redirect(url_for("hr.leave_view", lid=lid))
